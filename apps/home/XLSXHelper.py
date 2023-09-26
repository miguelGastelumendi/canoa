# -*- encoding: utf-8 -*-
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import apps.home.dbquery as dbquery
import re
import apps.home.XLSXHelperSQL as XLSXHelperSQL
from copy import copy
from string import ascii_uppercase
from .emailstuff import sendEmail
import apps.home.logHelper as logHelper

class XLSXHelper:
    def __init__(self, templateFileName: str, idProjeto: int, log: logHelper.Log):
        self.xlsx = openpyxl.load_workbook(templateFileName)
        self.idProjeto = idProjeto
        self.log = log
        self.projectData = self.getProjectData()
        self.XLSXfname = f"doc/xls_projectresults/ResultadosReflorestaSP_{self.projectData['username'].replace(' ', '_')}_" \
                f"{self.idProjeto}_{self.projectData['descProjeto'].replace(' ', '_')}.xlsx"

    colName = lambda self, name: ''.join([x for x in name if x in ascii_uppercase])

    # Funciona só para A1 até ZZ(n)
    def rowColIndex(self, cName: str) -> int:
        cName = cName.replace('$', '')
        v = [ord(c) - 64 for c in cName if c >= 'A' and c <= 'Z']
        return v[0] if len(v) == 1 else v[0] * 26 + v[1], int(cName[re.search(r"\d", cName).start():]) + 1 # first line is title

    def getProjectData(self):
        return dbquery.getDictFieldNamesValuesResultset(
            XLSXHelperSQL.SQLs['userNameDescProjeto'].format(self.idProjeto))

    def getDataAsRows(self, name: str):
        return dbquery.executeSQL(XLSXHelperSQL.SQLs[name].format(self.idProjeto))

    def getDataAsDict(self, name: str):
        return dbquery.getDictFieldNamesValuesResultset(XLSXHelperSQL.SQLs[name].format(self.idProjeto))

    def fillWorksheetFromRows(self, wsName: str, tbName: str, queryName=None):
        self.log.log(f'fillWorksheetFromRows:\n{locals()}')
        if queryName is None:
            queryName = wsName
        try:
            rows = self.getDataAsRows(queryName)
            ws = self.xlsx[wsName]
            if tbName not in self.xlsx.defined_names:  # so, it's table
                rangeCells = ws.tables[tbName].ref.split(':')
                initColNum, initRowNum = self.rowColIndex(rangeCells[0])
            else:
                rangeCells = [self.xlsx.defined_names[tbName].value.split('!')[1]]
                initColNum, initRowNum = self.rowColIndex(rangeCells[0])
                initRowNum = initRowNum - 1

            for i, row in enumerate(rows, initRowNum):
                for j, value in enumerate(row, initColNum):
                    if ws.cell(i, j).has_style:
                        ws.cell(i, j).number_format = copy(ws.cell(1, j).number_format)
                    ws.cell(i, j).value = value
                j += 1
                while ws.cell(2, j).value is not None:
                    ws.cell(i, j).value = ws.cell(2, j).value
                    if ws.cell(1, j).has_style:
                        ws.cell(i, j).number_format = copy(ws.cell(1, j).number_format)
                    j += 1
            if not tbName in self.xlsx.defined_names:
                # recreating Table in worksheet
                tab = Table(displayName=tbName,
                            ref=f"{rangeCells[0]}:{self.colName(rangeCells[1])}{i}",
                            tableStyleInfo=ws.tables[tbName].tableStyleInfo)
                del ws.tables[tbName]
                ws.tables.add(tab)
        except Exception as e:
            print("Erro: ")
            print("Tablename:", tbName)
            print()
            print(e)
            #raise Exception('Error.')

    def fillWorksheetFromDict(self, name: str):
        try:
            self.log.log(f'fillWorksheetFromRows:\n{locals()}')
            dict = self.getDataAsDict(name)
            for key, value in dict.items():
                worksheet, cell = self.xlsx.defined_names[key].value.split('!')
                self.xlsx[worksheet][cell] = value
        except Exception as e:
            self.log.log(f"Erro: em fillWorksheetFromDict\n{e}")

    def fillDistribution(self):
        pass

    def execute(self):
        self.fillWorksheetFromDict('Identificação')
        self.fillWorksheetFromRows('Silvicultura', 'tbSilv')
        self.fillWorksheetFromRows('Receitas', 'tbRec')
        self.fillWorksheetFromRows('Parametros', 'tbUnid', queryName='Unidades')
        self.fillWorksheetFromRows('Parametros', 'tbFaixa', queryName='Faixas')
        self.fillWorksheetFromRows('Parametros', 'TxDsc', queryName='txDsc')
        self.fillWorksheetFromRows('FluxoCaixaModelo', 'TxPoup', queryName='TxPoup')
        self.fillWorksheetFromRows('ResumoSilvicultura', 'tbResumo')
        self.fillWorksheetFromRows('FluxoCaixaFaixa', 'tbFcFaixa')
        self.fillWorksheetFromRows('Distribuição', 'tbDistribuicao', queryName='tbDistribuicao')
        self.fillDistribution()

    def save(self):
        print(f'Salvando {self.XLSXfname}...')
        self.xlsx.save(self.XLSXfname)

    def sendEMail(self):
        print(f"Enviando a planilha para {self.projectData['eMailEnvioResultado']}...")
        sendEmail(self.projectData['eMailEnvioResultado'],'emailSendSpreadsheet',
                  {'descProjeto': self.projectData['descProjeto']},
                  self.XLSXfname)

def GenerateXLSX(idProjeto: int, log: logHelper.Log):
    log.log('Generate XLSX')
    xlsHelper = XLSXHelper('doc/TemplatePlanilhaComProjetoDoUsuario.xlsx', idProjeto, log)
    xlsHelper.execute()
    xlsHelper.save()
    xlsHelper.sendEMail()